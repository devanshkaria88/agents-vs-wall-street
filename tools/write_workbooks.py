#!/usr/bin/env python
"""Write the four submission workbooks from a forecasts JSON file.

Deliberately dumb plumbing: no forecasting logic lives here. It copies each
supplied template, fills ONLY the three yellow forecast cells (Summary C7:C9),
cross-checks labels/units against challenge/companies.json, and refuses to
write anything non-numeric. See .claude/skills/output-validation.

Usage: .venv/bin/python tools/write_workbooks.py forecasts/forecasts.json

Input schema: {"HD-FY2026Q2.xlsx": {"Net sales": 45300.0, ...}, ...}
— keys are outputFile names, values map exact metric label -> number.
"""
import json
import math
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
YELLOW_ROWS = (7, 8, 9)  # Summary C7:C9, verified against all four templates


def fail(msg: str) -> None:
    print(f"FAIL {msg}", file=sys.stderr)
    sys.exit(1)


def main(forecasts_path: str) -> None:
    companies = json.loads((ROOT / "challenge" / "companies.json").read_text())
    forecasts = json.loads(Path(forecasts_path).read_text())
    (ROOT / "submission").mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).isoformat()

    for company in companies["companies"]:
        out_name = company["outputFile"]
        values = forecasts.get(out_name)
        if values is None:
            fail(f"{out_name}: no forecasts supplied — a blank cell scores 5.0")

        template = ROOT / "challenge" / "templates" / out_name
        target = ROOT / "submission" / out_name
        shutil.copy(template, target)

        wb = openpyxl.load_workbook(target)
        ws = wb["Summary"]
        for row, metric in zip(YELLOW_ROWS, company["metrics"]):
            label, units = metric["label"], metric["units"]
            if ws[f"A{row}"].value != label or ws[f"B{row}"].value != units:
                fail(f"{out_name}: template row {row} is {ws[f'A{row}'].value!r}/"
                     f"{ws[f'B{row}'].value!r}, expected {label!r}/{units!r}")
            if label not in values:
                fail(f"{out_name}: missing forecast for {label!r}")
            v = values[label]
            if not isinstance(v, (int, float)) or isinstance(v, bool) or not math.isfinite(v):
                fail(f"{out_name}: {label!r} = {v!r} is not a finite number")
            ws[f"C{row}"] = float(v)
            print(f"WROTE {out_name} C{row} {label} ({units}) = {float(v)} at {stamp}")
        wb.save(target)

    print("All four workbooks written.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        fail("usage: write_workbooks.py <forecasts.json>")
    main(sys.argv[1])
